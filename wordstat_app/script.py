import json
import os
import time
import logging
from django.conf import settings
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl

logging.config.dictConfig(settings.LOGGING)


def log_with_user_info(level, message, user_info='Unknown'):
    logger = logging.getLogger('django')
    logger.log(level, message, extra={'user_info': user_info})


month_dict = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def generate_yandex_cookies(user):
    """Генерация куки для авторизации на Яндекс.Вордстат."""
    return [
        {
            "name": "Session_id",
            "value": user.session_id,
            "domain": ".yandex.ru",
            "hostOnly": False,
            "path": "/",
            "secure": True,
            "httpOnly": True,
            "session": False,
            "firstPartyDomain": "",
            "partitionKey": None,
            "storeId": None,
        },
        {
            "name": "yandexuid",
            "value": user.yandexuid,
            "domain": ".yandex.ru",
            "hostOnly": False,
            "path": "/",
            "secure": True,
            "httpOnly": False,
            "session": False,
            "firstPartyDomain": "",
            "partitionKey": None,
            "storeId": None,
        },
    ]


def get_driver(user):
    """Инициализация и настройка веб-драйвера Firefox."""
    options = Options()
    options.add_argument('-headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    service = Service('/usr/local/bin/geckodriver')
    driver = webdriver.Firefox(service=service, options=options)

    driver.get("https://wordstat.yandex.ru/")
    cookies = generate_yandex_cookies(user)
    for cookie in cookies:
        driver.add_cookie(cookie)
    driver.get("https://wordstat.yandex.ru/")
    check_cookies(driver)
    return driver


def load_cookies(driver, cookies_file, user_info=None):
    """Загрузка куки в веб-драйвер."""
    logging.info(f"Загрузка куки из файла: {
                 cookies_file}", extra={'user_info': user_info})
    try:
        with open(cookies_file, "r") as f:
            cookies = json.load(f)
            for cookie in cookies:
                driver.add_cookie(cookie)
        logging.info("Куки загружены успешно.", extra={'user_info': user_info})
    except Exception as e:
        logging.error(f"Ошибка загрузки куки: {e}")


class InvalidCookiesError(Exception):
    """Исключение, вызываемое при недействительных куки."""
    pass


def check_cookies(driver, user_info=None):
    """Проверяет, авторизован ли пользователь в Яндекс.Wordstat."""
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "user-pic"))
        )
        logging.info("Пользователь успешно авторизован в Яндекс.Wordstat",
                     extra={'user_info': user_info})
        return True
    except Exception as e:
        logging.error(f"Ошибка авторизации в Яндекс.Wordstat: {
                      e}", extra={'user_info': user_info})
        raise InvalidCookiesError(
            "Куки недействительны или пользователь не авторизован"
        )


def fetch_wordstat_data(query, driver, retry_count=3, user_info=None):
    """Получение данных с Яндекс.Вордстат."""
    logging.info(f"Запрос данных для запроса: {
                 query}", extra={'user_info': user_info})
    for attempt in range(retry_count):
        try:
            base_url = (
                "https://wordstat.yandex.ru/?region=all&view=graph&words="
            )
            encoded_query = query.replace(" ", "%20")
            full_url = base_url + encoded_query
            driver.get(full_url)

            WebDriverWait(driver, 40).until(
                EC.presence_of_element_located(
                    (By.CLASS_NAME, "table__wrapper"))
            )
            logging.info("Данные получены успешно.",
                         extra={'user_info': user_info})
            return driver.page_source
        except Exception as e:
            logging.error(
                f"Попытка {attempt + 1}/{retry_count}: "
                f"Ошибка получения данных для запроса '{query}' - {e}",
                extra={'user_info': user_info}
            )
            time.sleep(10)
    raise Exception(
        f"Не удалось получить данные для запроса '{
            query}' после {retry_count} попыток"
    )


def parse_wordstat_html(html, start_year, end_year, user_info=None):
    """Парсинг HTML-страницы с данными Яндекс.Вордстат."""
    logging.info(
        f"Парсинг HTML-страницы для периода: {start_year} - {end_year}",
        extra={'user_info': user_info})
    soup = BeautifulSoup(html, "html.parser")
    results = []

    table = soup.find("table", {"class": "table__wrapper"})
    if not table:
        raise Exception("Данные не найдены")

    rows = table.find_all("tr")
    for row in rows:
        month_cell = row.find("td", {"class": "table__content-cell"})
        count_cell = row.find("td", {"class": "table__level-cell"})
        if month_cell and count_cell:
            month = month_cell.text.strip()
            if start_year <= int(month.split()[-1]) <= end_year:
                count = int(count_cell.text.strip().replace(" ", ""))
                results.append((month, count))
    logging.info("Данные успешно спаршены.", extra={'user_info': user_info})
    return results


def process_without_year(
    results, start_month, end_month,
    start_year, end_year, threshold=0.05, user_info=None
):
    """Обработка результатов без учета года."""
    logging.info(
        f"Обработка результатов без учета года: {
            start_month} - {end_month}, {start_year} - {end_year}",
        extra={'user_info': user_info}
    )
    filtered_results = []
    for month, count in results:
        month_name, year = month.split()
        month_num = month_dict[month_name.lower()]
        year = int(year)
        if (year == start_year and month_num >= start_month) or (
            year == end_year and month_num <= end_month
        ):
            filtered_results.append((month_name, count))
    return find_peak_months(filtered_results, threshold)


def need_year(start_year, start_month, end_year, end_month, user_info=None):
    """Проверка, нужно ли включать год в результаты."""
    logging.info(
        f"Проверка необходимости включения года в результаты: {
            start_year}, {start_month}, {end_year}, {end_month}",
        extra={'user_info': user_info}  # Добавь  extra
    )
    if start_year == end_year:
        return start_month > end_month
    else:
        return True


def process_with_year(
    results, start_month, end_month,
    start_year, end_year, threshold=0.05, user_info=None
):
    """Обработка результатов с учетом года."""
    logging.info(
        f"Обработка результатов с учетом года: {
            start_month} - {end_month}, {start_year} - {end_year}",
        extra={'user_info': user_info}
    )
    filtered_results = []
    for month, count in results:
        month_name, year = month.split()
        month_num = month_dict[month_name.lower()]
        year = int(year)
        if start_year <= year <= end_year:
            if (
                (year == start_year and month_num >= start_month)
                or (year == end_year and month_num <= end_month)
                or (start_year < year < end_year)
            ):
                filtered_results.append((f"{month_name} {year}", count))
    return find_peak_months(filtered_results, threshold)


def find_peak_months(filtered_results, threshold, user_info=None):
    """Поиск пиковых месяцев."""
    logging.info(f"Поиск пиковых месяцев: {
                 threshold}", extra={'user_info': user_info})
    if not filtered_results:
        logging.info("Пиковые месяцы не найдены.",
                     extra={'user_info': user_info})
        return [], []
    max_count = max(count for _, count in filtered_results)
    max_months = [month for month,
                  count in filtered_results if count == max_count]
    near_max_months = [
        month
        for month, count in filtered_results
        if (max_count - count) / max_count <= threshold and count != max_count
    ]
    logging.info(f"Пиковые месяцы найдены: {max_months}, {
                 near_max_months}", extra={'user_info': user_info})
    return max_months, near_max_months


def read_excel_data(file_path, sheet_name, user_info=None):
    """Чтение данных из Excel-файла."""
    logging.info(f"Чтение данных из файла: {file_path}, лист: {
                 sheet_name}", extra={'user_info': user_info})
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(
            min_row=2,
            max_row=sheet.max_row,
            values_only=True
        ):
            data.append(row)
        logging.info("Данные из Excel успешно считаны.",
                     extra={'user_info': user_info})
        return data, workbook, sheet
    except KeyError:
        logging.error(f"{sheet_name} не найден в Excel-файле.",
                      extra={'user_info': user_info})
        raise ValueError(f"{sheet_name} не найден в Excel-файле.",
                         extra={'user_info': user_info})
    except Exception as e:
        logging.error(f"Ошибка чтения данных из Excel: {
                      e}", extra={'user_info': user_info})
        raise


def update_excel(sheet, row_index, result_column,
                 value, excel_file_path, user_info=None):
    """Обновление данных в Excel-файле."""
    logging.info(f"Обновление Excel: строка {row_index}, значение '{
                 value}'", extra={'user_info': user_info})
    try:
        sheet.cell(row=row_index, column=result_column, value=value)
        workbook = sheet.parent
        workbook.save(excel_file_path)
        workbook.close()
        logging.info(
            f"Файл Excel сохранен и закрыт после обновления строки {
                row_index}",
            extra={'user_info': user_info}
        )
    except Exception as e:
        logging.error(f"Ошибка обновления Excel: {e}")


def process_excel_file(
    file_path,
    sheet_name,
    data_column,
    result_column,
    start_cell,
    start_year,
    end_year,
    start_month,
    end_month,
    user,
    user_info=None
):
    """Обрабатывает Excel-файл, извлекая данные и обновляя файл результатами"""

    upload_dir = os.path.dirname(file_path)
    processed_dir = settings.PROCESSED_FILES_DIR
    os.makedirs(upload_dir, exist_ok=True)
    os.makedirs(processed_dir, exist_ok=True)

    try:
        data, workbook, sheet = read_excel_data(
            file_path, sheet_name, user_info=user_info)
    except ValueError as e:
        return None, False, str(e)

    try:
        driver = get_driver(user)
    except InvalidCookiesError:
        logging.error("Недействительные куки или ошибка авторизации", extra={
                      'user_info': user_info})
        return None, False, "Недействительные куки или ошибка авторизации"

    try:
        for row_index in range(start_cell - 1, len(data)):
            row = data[row_index]
            query = row[data_column - 1]
            if query is None:
                logging.info(
                    f"Пустая ячейка, остановка обработки на строке "
                    f"{row_index + 2}", extra={'user_info': user_info}
                )
                break

            query = str(query)

            logging.info(f"Обработка запроса: {query}", extra={
                         'user_info': user_info})
            print(f"Processing query {row_index + 1}/{len(data)}: {query}")

            try:
                html = fetch_wordstat_data(query, driver, user_info=user_info)
                results = parse_wordstat_html(
                    html, start_year, end_year, user_info=user_info)
                if results:
                    if need_year(start_year, start_month, end_year,
                                 end_month, user_info=user_info):
                        max_months, near_max_months = process_with_year(
                            results, start_month, end_month,
                            start_year, end_year, user_info=user_info
                        )
                    else:
                        max_months, near_max_months = process_without_year(
                            results, start_month, end_month,
                            start_year, end_year, user_info=user_info
                        )

                    output_str = f"{', '.join(max_months)}|{
                        ', '.join(near_max_months)}"
                else:
                    output_str = "No data found"
            except Exception as e:
                output_str = f"Error processing query '{query}': {e}"

            update_excel(sheet, row_index + 2, result_column,
                         output_str, file_path, user_info=user_info)
    finally:
        driver.quit()

    processed_file_path = os.path.join(
        processed_dir, f"processed_{os.path.basename(file_path)}"
    )
    workbook.save(processed_file_path)
    workbook.close()
    return (processed_file_path, True, None)
